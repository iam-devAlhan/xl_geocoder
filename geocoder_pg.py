# -*- coding: utf-8 -*-
#-----------------------------------------------------------------------------------------
# Nazwa:     geocoder_pg
# Opis:      Geokoduje adresy i zwraca warstwę shp na podstawie listy zapisanej w xlsx. 
#
# Autor:     Przemek Garasz
# Data utw:  2018-02-20
# Wersja:    1.0
#-----------------------------------------------------------------------------------------

import os
import geocoder
import shapefile
import datetime
import re
from time import sleep
from requests import Session
from openpyxl import load_workbook, Workbook

class FakeGC:
    '''Sztuczna klasa do symulacji statusów geocodera.osm'''

    def __init__(self, ok, status, status_code=-999, timeout=-999):
        self.ok = ok
        self.status = status
        self.status_code = status_code
        self.timeout = timeout


def create_empty_shp(path, field_params_list, shapeType):
    '''
    Tworzy pusty plik shp ze strukturą atrubutów
    shp_writer - instancja klasy Writer modułu shapefile
    field_params_list - parametry pola tabeli w postaci listy list [nazwa, typ, rozmiar]
    '''
    with shapefile.Writer(path, shapeType) as shp:
        for field_params in field_params_list:
            shp.field(*field_params)


def add_fields_to_shp(shp_writer, field_params_list):
    '''
    Dodaje atrybuty do shp
    shp_writer - instancja klasy Writer modułu shapefile
    field_params_list - parametry atrybutu w postaci listy list 
                        [nazwa, typ, rozmiar]
    '''
    for field_params in field_params_list:
        shp_writer.field(*field_params)


def sanitize_n_unicode(string):
    try:
        if string:
            return unicode(string).strip()
        else:
            return 'BRAK DANYCH'
    except AttributeError:
        return 'ZŁY TYP DANYCH'


def parse_street_name(street_name, name_filter=None, remove_abbreviation=False,
                      building_number_first=False):
    '''
    Przetwarza i filtruje nazwę ulicy. 
    (Opcja) Zupełnie odrzuca nazwę jeżeli zawiera ciąg tekstu z listy "filter". 
    (Opcja) Usuwa skróty zakończone kropką (ul. Gen. Św. M.).
    (Opcja) Znajduje numer budynku na końcu i przenosi na początek 
            (rozwiązanie pod osm).
    '''
    regex = re.compile('\w+\.', re.UNICODE)  # Szuka słów zakończonych "."
    regex2 = re.compile(ur'(?<= )\d*((?<=\d)(/|\\))?\d+[a-zA-Z]?$', re.UNICODE)  # Szuka numeru budynku na końcu ciągu znaków

    if name_filter:
        for substring in name_filter:
            if substring in street_name:
                return False
    if remove_abbreviation:
        street_name = re.sub(regex, '', street_name).strip()
    if building_number_first:
        try:
            building_number = re.search(regex2, street_name).group()
            street_name = building_number + ', ' + street_name.replace(building_number, '')
        except AttributeError:
            None
    return unicode(street_name.strip())


if __name__ == "__main__":

### Konfiguracja --------------------------------------------------------------
    
    xls_path = 'dane\\Instalacje_PZ_30092018.xlsx'  # zrodlo danych do geokodowania
    xls_name = os.path.splitext(os.path.basename(xls_path))[0]
    xls_min_row = 2
    xls_max_row = 10
    xls_max_column = 5
    illegal_street_name_substrings = [u'dz.', u'ew.', u' działki ', u' nr ', u' obręb ', u' ewid ']
    
    now = datetime.datetime.now()
    timestamp = now.strftime('%Y-%m-%d_%H-%M-%S')
    output_dir = 'output_' + timestamp
    
    output_shp_name = xls_name  # moduł shapefile ignoruje rozszerzenia plików
    output_shp_path = os.path.join(output_dir, output_shp_name)
    
    incorrect_data_xls_name = 'NIEPOPRAWNE_ADRESY_' + xls_name + '.xlsx'
    incorrect_data_xls_path = os.path.join(output_dir, incorrect_data_xls_name)

    delay = 1.2  # opóźnienie zapytania do serwera w sekundach 

    # Konfiguracja atrybutów shp
    fields_config = [
        ['NAZWA', 'C', 255],
        ['UL_NR_ORG', 'C', 255],
        ['UL_NR_MOD', 'C', 255],
        ['KOD', 'C', 255],
        ['MIEJSC', 'C', 255],
        ['WOJ', 'C', 255],
        ['OSM', 'C', 255]
    ]

### Instrukcje ----------------------------------------------------------------

    # Odczyt xls z danymi
    wb = load_workbook(xls_path, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(min_row=xls_min_row, max_row=xls_max_row,
                        max_col=xls_max_column, values_only=True)
    # Xls na błędne adresy
    incorrect_data_wb = Workbook()
    incorrect_data_ws = incorrect_data_wb.active
    incorrect_data_ws.title = 'No result'
    incorrect_data_ws.append(['Nazwa', 'ul_nr_org', 'ul_nr', 'kod', 'miejscowosc', 'woj',
                             'nr_wiersza', 'gc_status', 'gc_status_code', 'gc_timeout'])


    print '\n' + 'GEOKODOWANIE - START' + '\n'

    with shapefile.Writer(output_shp_path, 1) as shp:

        add_fields_to_shp(shp, fields_config)
        
        with Session() as session:

            for i, row in enumerate(rows):
                
                print i + xls_min_row

                # Odczyt danych z xls
                nazwa = sanitize_n_unicode(row[0])         # A - nazwa
                ul_nr_org = sanitize_n_unicode(row[1])     # B - ulica + numer
                kod = sanitize_n_unicode(row[2])           # C - kod pocztowy
                miejsc = sanitize_n_unicode(row[3])        # D - miejscowosc
                woj = sanitize_n_unicode(row[4])           # E - wojewodztwo

                print u'      dane: {0}'.format(ul_nr_org)

                ul_nr = parse_street_name(street_name=ul_nr_org, name_filter=illegal_street_name_substrings,
                            remove_abbreviation=True, building_number_first=True)
                
                if ul_nr:
                    adres = ul_nr.strip() + ', ' + kod + ' ' + miejsc + ', ' + 'Polska'
                    print u'   szukane: {0}'.format(adres)
                    gc = geocoder.osm(adres, session=session)
                else:
                    gc = FakeGC(False, u"BŁĄD - NIERPAWIDŁOWA NAZWA ULICY")

                if ul_nr == ul_nr_org:
                    ul_nr = ''

                if gc.ok:
                    shp.point(gc.lng, gc.lat)
                    shp.record(nazwa, ul_nr_org, ul_nr, kod, miejsc, woj, gc.osm)
                    print '       lat: {0}; lng: {1}'.format(gc.lat, gc.lng)
                else:
                    print u'       {0} (status:{1}, timeout:{2})'.format(
                        gc.status, gc.status_code, gc.timeout)
                    incorrect_data_ws.append(
                        [nazwa, ul_nr_org, ul_nr, kod, miejsc, woj,
                         i+1, gc.status, gc.status_code, gc.timeout])
                    try:
                        incorrect_data_wb.save(incorrect_data_xls_path)
                    except:
                        incorrect_data_wb.save(
                            incorrect_data_xls_path.replace(xls_name, xls_name + '_alt'))

                sleep(delay)  # przeciwdziała banowaniu
