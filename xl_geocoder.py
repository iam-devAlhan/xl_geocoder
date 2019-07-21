"""
 -----------------------------------------------------------------------------------------
 Name:     xl_geocoder
 Desc:     Geocodes addresses from Excel spreadsheets and saves the results as shp files.

 Author:   Przemek Garasz
 Created:  2018-02-20
 Modified: 2018-07-21
 Wersja:   1.7
 ----------------------------------------------------------------------------------------
"""

import os
import geocoder
import shapefile
import datetime
import re
from time import sleep
from requests import Session
from openpyxl import load_workbook, Workbook
from tools import load_config
from tools.xl import get_fields_properties_from_worksheet
from tools.shp import add_fields_to_shp, create_prj_file


class FakeGC:
    """Simulates geocoder.osm output status"""

    def __init__(self, ok, status, status_code=-999, timeout=-999, osm=''):
        self.ok = ok
        self.status = status
        self.status_code = status_code
        self.timeout = timeout
        self.osm = osm


def sanitize_value(value, replace_none=False):
    try:
        if replace_none:
            if not value:
                return 'NO DATA'
        if value:
            return str(value).strip()
        else:
            return ''
    except TypeError:
        return 'INCORRECT TYPE'


def parse_street_name(street_name, name_filter=None, expand_abbrev=None,
                      remove_abbrev=False, building_number_first=False):
    """Street name parsing and filtering

    Args:
        name_filter   - (list) - returns False if string from the list is found in the street name
        expand_abbrev - (dict) - if key is found in the street name it will be replaced by its value
        remove_abbrev - (bool) - removes every word ending with a '.'
        building_number_first (bool) - moves the building number from the end of the string to the beginning

    Returns:
        string, False or None
    """
    regex = re.compile(r'\w+\.', re.UNICODE)
    regex2 = re.compile(r'((?<= )\d*)((?<=\d)/|(?<=\d)-|(?<=\d)\\)?(\d+)(?: |/|\\)?([a-zA-Z])?$')

    if name_filter:
        for substring in name_filter:
            if substring in street_name:
                return False
    if expand_abbrev:
        for key in expand_abbrev:
            street_name = re.sub(key, expand_abbrev[key], street_name, flags=re.IGNORECASE)
    if remove_abbrev:
        street_name = re.sub(regex, '', street_name).strip()
    if building_number_first:
        try:
            match = re.search(regex2, street_name)
            building_number = match.group()
            mod_number = match.expand(r'\1\2\3\4')
            street_name = mod_number + ', ' + street_name.replace(building_number, '').strip()
        except AttributeError:
            None
    return street_name.strip()


if __name__ == "__main__":

    # Config ----------------------------------------------------------------------------

    config = load_config('config.yaml')

    xls_path = config['xls']['path']
    xls_name = os.path.splitext(os.path.basename(xls_path))[0]
    xls_has_header = config['xls']['has_header']
    xls_min_row = config['xls']['min_row']
    xls_max_row = config['xls']['max_row']
    xls_max_column = config['xls']['max_column']

    address_columns_indxs = config['address']['col_indxs']
    illegal_street_names = config['address']['illegal_street_names']
    abbrev_dict = config['address']['abbrev_expansions']
    remove_abbrev = config['address']['remove_abbrev']

    strict_search = config['strict_search']

    now = datetime.datetime.now()
    timestamp = now.strftime('%Y-%m-%d_%H-%M-%S')
    output_dir = 'output_' + timestamp

    output_shp_name = xls_name  # shapefile package ignores file extensions
    output_shp_path = os.path.join(output_dir, output_shp_name)

    no_results_xls_name = 'NO_RESULTS_' + xls_name + '.xlsx'
    no_results_xls_path = os.path.join(output_dir, no_results_xls_name)

    delay = 1.2  # OSM query delay, >1s gets you banned

    additional_shp_fields = [
        ['QUERY', 'C', 255],
        ['OSM_ANSW', 'C', 255],
        ['CONFIDENCE', 'F', 5, 2]
    ]

    # Main -------------------------------------------------------------------------------

    wb = load_workbook(xls_path, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(min_row=xls_min_row, max_row=xls_max_row,
                        max_col=xls_max_column, values_only=True)

    fields_config = get_fields_properties_from_worksheet(ws, xls_has_header)
    shp_fields_config = fields_config + additional_shp_fields

    no_results_wb = Workbook()
    no_results_ws = no_results_wb.active
    no_results_ws.title = 'No result'
    if xls_has_header:
        column_headers = [field_property[0] for field_property in fields_config]
    else:
        column_headers = ['' for field_property in fields_config]
    no_results_ws.append(column_headers + ['query', 'gc_status', 'gc_status_code', 'gc_timeout'])


    print('\n' + 'GEOCODING...' + '\n')

    with shapefile.Writer(output_shp_path, 1) as shp:

        add_fields_to_shp(shp, shp_fields_config)
        create_prj_file(output_shp_path + '.prj', 4326, 'GCS_WGS_1984')

        with Session() as session:

            for i, row in enumerate(rows):

                print(i + xls_min_row)

                idx = address_columns_indxs
                st_name_num            = sanitize_value(row[idx['st_name_num']])
                secondary_place_name   = sanitize_value(row[idx['secondary_place_name']])
                postal_code            = sanitize_value(row[idx['postal_code']])  # currently not in use
                primary_place_name     = sanitize_value(row[idx['primary_place_name']])
                county                 = sanitize_value(row[idx['county']])
                province               = sanitize_value(row[idx['province']])  # currently not in use


                if secondary_place_name != '' and st_name_num != '':     # named streets, shared postal code - villages
                    print(f'      data: {st_name_num}, {secondary_place_name}')
                    parsed_st_name = parse_street_name(street_name=st_name_num, name_filter=illegal_street_names,
                                                  expand_abbrev=abbrev_dict, remove_abbrev=remove_abbrev, building_number_first=True)
                    address = parsed_st_name + ', ' + secondary_place_name + ', ' + county

                elif secondary_place_name != '':         # no street names (just building numbers) - small villages and other settlements
                    print(f'      data: {secondary_place_name}')
                    parsed_st_name = parse_street_name(street_name=secondary_place_name, name_filter=illegal_street_names,
                                                  building_number_first=True)
                    address = parsed_st_name + ', ' + county

                elif primary_place_name != '' and st_name_num != '':         # named streets and own postal code - large villages, towns, cities
                    print(f'      data: {st_name_num}, {primary_place_name}')
                    parsed_st_name = parse_street_name(street_name=st_name_num, name_filter=illegal_street_names,
                                                  expand_abbrev=abbrev_dict, remove_abbrev=remove_abbrev, building_number_first=True)
                    if primary_place_name.lower() == county.lower():
                        address = parsed_st_name + ', ' + primary_place_name
                    else:
                        address = parsed_st_name + ', ' + primary_place_name + ', ' + county
                else:
                    address = None


                if strict_search:
                    has_leading_number = re.match(r'^\d+\S*', address)  # Correct address must contain a building number

                    if address and has_leading_number:
                        gc = geocoder.osm(address, session=session)
                    else:
                        gc = FakeGC(False, u"ERROR - INCORRECT ADDRESS")
                else:
                    if address:
                        print(f'     query: {address}')
                        gc = geocoder.osm(address, session=session)

                        while 'No results' in gc.status and address.find(',') > -1:  #  if no result
                            i = address.find(',')
                            if i > -1:
                                address = address[i+1:]  # drop the part before the comma
                                sleep(delay)
                                print(f'     query: {address}')
                                gc = geocoder.osm(address, session=session)
                    else:
                        gc = FakeGC(False, u"ERROR - INCORRECT ADDRESS")


                if gc.ok:
                    confidence = gc.current_result.confidence
                    shp.point(gc.lng, gc.lat)
                    shp.record(*row, address, gc.osm, confidence)
                    print(f'    result: LAT {gc.lat}; LNG {gc.lng}; confidence: {confidence}')
                else:
                    print(f'       {gc.status} (status:{gc.status_code}, timeout:{gc.timeout})')
                    no_results_ws.append(row + (address, gc.status, gc.status_code, gc.timeout))
                    try:
                        no_results_wb.save(no_results_xls_path)
                    except Exception:
                        no_results_wb.save(
                            no_results_xls_path.replace(xls_name, xls_name + '_alt'))

                sleep(delay)
