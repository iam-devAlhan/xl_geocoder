from datetime import datetime
from openpyxl import load_workbook, Workbook


def get_column_samples_from_worksheet(worksheet_object, has_header=None, row_number=None):
    """Returns two element lists with column name and value from the chosen row

    Args:
        worksheet_object - openpyxl worksheet object
        has_header (bool, optional) -
            True  - column names based on the first row
            False - column names as alphabet letters
        row_number (int, optional) - row number to take samples from, 
                    if none provided first row will be used (or second if first is header)

    Returns:
        list: Two element lists with column name and value from the chosen row
    """

    ws = worksheet_object
    max_row = ws.max_row

    if row_number:
        if 1 <= row_number <= max_row:
            i = row_number
        else:
            raise IndexError('Row number out of range. Max row number: ' + str(max_row))
    else:
        if has_header:
            i = 2
        else:
            i = 1

    if has_header:
        column_names = [str(cell.internal_value) for cell in ws[1]]
        column_types = [cell.internal_value for cell in ws[i]]
    else:
        column_names = [cell.column_letter for cell in ws[1]]
        column_types = [cell.internal_value for cell in ws[i]]

    return list(map(list, zip(column_names, column_types)))


def _validate_custom_properties(custom_properties):
    valid = []

    case1 = custom_properties == 'auto'
    case2 = isinstance(custom_properties, dict)

    if not (case1 or case2):
        raise TypeError('Incorrect custom_properties value')

    for data_type in custom_properties:
        if data_type in [str, int, float, bool, datetime] or data_type is None:
            p0, p1, p2 = custom_properties[data_type]
            p0_valid = p0 in ['C', 'N', 'F', 'L', 'D'] or p0 is None
            p1_valid = isinstance(p1, int)
            p2_valid = isinstance(p2, int)
            valid.append(p0_valid and p1_valid and p2_valid)
        else:
            TypeError('Incorrect data type')

        if not all(valid):
            raise ValueError(f'Incorrect config for: {data_type}')


def determine_field_properties(value, custom_properties=None):
    """Determines value properties in shp table context (type, length, precision)
    Return lists with fields config used for shapefile creation

    Args:
        value - analyzed value
        custom_properties (optional) -
            Konfiguracja parametrów typów pola tabeli
              None - auto determine value type,
                     length and precision from `default_properties`
            "auto" - auto determine value type, length and precision for floats
              dict - auto determine value type,
                     override `default_properties` length and precision with values
                     from the provided dict, e.g.
                     {str: ['C', 255, 0]}

    Returns:
        list - Lists of field properties to be used by the shapefile package
    """

    default_properties = {
        # [type_letter, field_length, decimal_places]
        str:      ['C', 255, 0],
        int:      ['N', 9, 0],
        float:    ['N', 6, 2],
        bool:     ['L'],
        datetime: ['D'],
        None:     ['C', 255, 0]
    }

    value_type = type(value)

    if custom_properties == 'auto':
        if value_type == str:
            length = len(value)
            default_properties[str][1] = length
        if value_type == int:
            length = len(str(value))
            default_properties[int][1] = length
        if value_type == float:
            s = str(value)
            length = len(s.split('.')[0])
            decimal = len(s.split('.')[1])
            default_properties[float][1] = length + decimal
            default_properties[float][2] = decimal
        field_types_properties = default_properties
    elif custom_properties:
        default_properties.update(custom_properties)

    field_types_properties = default_properties

    if value_type in [str, int, float, bool, datetime]:
        field_property = field_types_properties[value_type]
    elif not value:
        field_property = field_types_properties[None]
    else:
        raise TypeError('Incorrect data type')

    return list(field_property)


def get_fields_properties_from_worksheet(worksheet_object, has_header=None, row_number=None,
                                        custom_properties=None):
    """Analyzes worksheet content and returns corresponding shapefile table config

      Args:
        worksheet_object - openpyxl worksheet object
        has_header (bool, optional) -
            True  - column names based on the first row
            False - column names as alphabet letters
        row_number (int, optional) - row number to take samples from, 
                    if none provided first row will be used (or second if first is header)
        custom_properties (optional) -
            Konfiguracja parametrów typów pola tabeli
              None - auto determine value type,
                     length and precision from `default_properties`
            "auto" - auto determine value type, length and precision for floats
              dict - auto determine value type,
                     override `default_properties` length and precision with values
                     from the provided dict, e.g.
                     {str: ['C', 255, 0]}
    Returns:
        list - list of lists - ['field_name', 'data_type_symbol', 'length', 'decimal']
    """
    column_samples = get_column_samples_from_worksheet(worksheet_object, has_header, row_number)

    if custom_properties:
        _validate_custom_properties(custom_properties)

    properties = []

    for sample in column_samples:
        name, value = sample
        field_properties = determine_field_properties(value, custom_properties)
        field_properties.insert(0, name)
        properties.append(field_properties)

    return properties
