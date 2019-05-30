from datetime import datetime
from openpyxl import load_workbook, Workbook


def get_column_samples_from_worksheet(worksheet_object, has_header=None, row_number=None):
    """Zwraca dwuelementowe listy z nazwą kolumny i wartością wybranego wiersza

    Args:
        worksheet_object - obiekt skoroszytu openpyxl
        has_header (bool, optional) -
            True  - nazwy kolumn będą brane z pierwszego wiersza tabeli
            False - nazwy kolumn to kolejne litery alfabetu
        row_number (int, optional) -
                    Numer wiersza, z którego będę brane wartości do analizy.
                    Domyślnie brany jest pierwszy lub drugi wiersz tabeli

    Returns:
        list: Dwuelementowa lista z nazwą kolumny i wartością z wybranego wiersza
    """

    ws = worksheet_object
    max_row = ws.max_row

    if row_number:
        if 1 <= row_number <= max_row:
            i = row_number
        else:
            raise IndexError('Numer wiersza poza zakresem. Maks. numer wiersza to' + str(max_row))
    else:
        if has_header:
            i = 2
        else:
            i = 1

    if has_header:
        column_names = [str(cell.internal_value) for cell in ws[1]]
        column_types = [cell.internal_value for cell in ws[i]]
    else:
        column_names = [cell.column_letter for cell in ws[1]]
        column_types = [cell.internal_value for cell in ws[i]]

    return list(map(list, zip(column_names, column_types)))


def _validate_custom_properties(custom_properties):
    valid = []

    case1 = custom_properties == 'auto'
    case2 = isinstance(custom_properties, dict)

    if not (case1 or case2):
        raise TypeError('Nieprawidłowa wartość argumentu custom_properties')

    for data_type in custom_properties:
        if data_type in [str, int, float, bool, datetime] or data_type is None:
            p0, p1, p2 = custom_properties[data_type]
            p0_valid = p0 in ['C', 'N', 'F', 'L', 'D'] or p0 is None
            p1_valid = isinstance(p1, int)
            p2_valid = isinstance(p2, int)
            valid.append(p0_valid and p1_valid and p2_valid)
        else:
            TypeError('Nieobsługiwany typ danych')

        if not all(valid):
            raise ValueError(f'Błędna konfiguracja dla {data_type}')


def determine_field_properties(value, custom_properties=None):
    """Rozpoznaje typ przekazanego argumentu i zwraca listę z konfiguracją
    pola tabeli do zastosowanie w module shapefile

    Args:
        value - wartość do analizy
        custom_properties (optional) -
            Konfiguracja parametrów typów pola tabeli
              None - typ dobierany jest automatycznie, pozostałe
                     właściwości według `default_properties`
            "auto" - automatycznie dobiera typ, długość
                     i ilość miejsc dziesiętnych (float)
              dict - typ dobierany automatycznie,
                     pozwala na ręczne nadpisanie wskazanych
                     wartości `default_properties`
                      {str: ['C', 255, 0]}

    Returns:
        list - lista właściwości pól tabeli
               do zastosowanie w module shapefile
    """

    default_properties = {
        # [type_letter, field_length, decimal_places]
        str:      ['C', 255, 0],
        int:      ['N', 9, 0],
        float:    ['N', 6, 2],
        bool:     ['L'],
        datetime: ['D'],
        None:     ['C', 255, 0]
    }

    value_type = type(value)

    if custom_properties == 'auto':
        if value_type == str:
            length = len(value)
            default_properties[str][1] = length
        if value_type == int:
            length = len(str(value))
            default_properties[int][1] = length
        if value_type == float:
            s = str(value)
            length = len(s.split('.')[0])
            decimal = len(s.split('.')[1])
            default_properties[float][1] = length + decimal
            default_properties[float][2] = decimal
        field_types_properties = default_properties
    elif custom_properties:
        default_properties.update(custom_properties)

    field_types_properties = default_properties

    if value_type in [str, int, float, bool, datetime]:
        field_property = field_types_properties[value_type]
    elif not value:
        field_property = field_types_properties[None]
    else:
        raise TypeError('Nieobsługiwany typ danych')

    return list(field_property)


def get_fields_properties_from_workseet(worksheet_object, has_header=None, row_number=None,
                                        custom_properties=None):
    """Zwraca listę z konfiguracją pól tabeli atrybutów
    do wykorzystania przy tworzeniu plików shp za pomocą modułu shapefiles

      Args:
        worksheet_object - obiekt skoroszytu openpyxl
        has_header (bool, optional) -
            True  - nazwy kolumn będą brane z pierwszego wiersza tabeli
            False - nazwy kolumn to kolejne litery alfabetu
        row_number (int, optional) -
                    Numer wiersza, z którego będę brane wartości do analizy.
                    Domyślnie brany jest pierwszy lub drugi wiersz tabeli
        custom_properties (optional) -
            Konfiguracja parametrów typów pola tabeli
                None  - typ dobierany jest automatycznie, pozostałe
                        właściwości według `default_properties`
               "auto" - automatycznie dobiera typ, długość
                        i ilość miejsc dziesiętnych (float)
                dict  - typ dobierany automatycznie, pozwala na ręczne nadpisanie
                        wybranych wartości `default_properties`,
                        podawać słowmnik o postaci:
                        {str: ['C', 255, 0]}
    Returns:
        list - list of lists - ['field_name', 'data_type_symbol', 'length', 'decima']
    """
    column_samples = get_column_samples_from_worksheet(worksheet_object, has_header, row_number)

    if custom_properties:
        _validate_custom_properties(custom_properties)

    properties = []

    for sample in column_samples:
        name, value = sample
        field_properties = determine_field_properties(value, custom_properties)
        field_properties.insert(0, name)
        properties.append(field_properties)

    return properties


if __name__ == "__main__":
    from xl_geocoder import create_empty_shp

    xls_path = r'demo_data\DPSiPOC.xlsx'
    xls_path = r'demo_data\a.xlsx'
    shp_name = 'test'

    wb = load_workbook(xls_path, read_only=True)
    ws = wb.active

    custom_properties = None
    # custom_properties = 'auto'
    # custom_properties = {str: ['C', 100, 0],
    #                      float: ['N', 10, 8],
    #                      int: ['N', 9, 0]}

    fields_properties = get_fields_properties_from_workseet(ws,
                                has_header=True,
                                custom_properties=custom_properties,
                                row_number=3)

    # for p in fields_properties:
    #     print(p)

    create_empty_shp(shp_name, fields_properties, 1)
